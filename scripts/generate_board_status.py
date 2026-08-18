#!/usr/bin/env python3
"""Y클라우드 게시판 현황판 생성기.

원본 엑셀(콘텐츠 현황 / 변경이력 시트)을 읽어 docs/board-status.html을 재생성한다.
사용법: python generate_board_status.py <엑셀경로> [출력경로]
"""
import sys
import json
import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = REPO_ROOT / "docs" / "board-status.template.html"
DEFAULT_OUTPUT = REPO_ROOT / "docs" / "board-status.html"

STATUS_MAP = {"게시중": "live", "준비중": "prep"}
HIST_TYPE_MAP = {"신규생성": "create", "수정": "edit", "삭제": "delete"}


def to_iso_date(value):
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if value in (None, "", "확인 필요"):
        return None
    return str(value).strip()


def load_items(wb):
    ws = wb["콘텐츠 현황"]
    items = {}
    order = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        item_id, level, subject, grade, board, name, itype, count, status, last_update = row[:10]
        items[item_id] = {
            "id": item_id,
            "level": level,
            "subject": subject,
            "grade": grade,
            "board": board,
            "name": name,
            "type": itype,
            "count": count if isinstance(count, (int, float)) else None,
            "status": STATUS_MAP.get(status, "prep"),
            "last_update": to_iso_date(last_update),
            "history": [],
        }
        order.append(item_id)
    return items, order


def load_history(wb, items):
    ws = wb["변경이력"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        item_id, date, kind, action, who = row[:5]
        if item_id not in items:
            continue
        items[item_id]["history"].append({
            "date": to_iso_date(date),
            "type": HIST_TYPE_MAP.get(kind, "edit"),
            "action": action,
            "who": who,
        })


def build_data(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    items, order = load_items(wb)
    load_history(wb, items)
    return [items[i] for i in order]


def render(xlsx_path, output_path, source_label=None):
    data = build_data(xlsx_path)
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    meta = {
        "generated_at": f"{now} (KST)",
        "source_name": source_label or Path(xlsx_path).name,
    }

    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    html = template.replace(
        "/*__BOARD_DATA_JSON__*/[]", json.dumps(data, ensure_ascii=False)
    ).replace(
        "/*__BOARD_META_JSON__*/{}", json.dumps(meta, ensure_ascii=False)
    )

    output_path = Path(output_path)
    old = output_path.read_text(encoding="utf-8") if output_path.exists() else None
    output_path.write_text(html, encoding="utf-8", newline="\n")

    # generated_at은 실행할 때마다 달라지므로, 그 줄을 제외하고 실제 데이터 변경 여부를 판단한다.
    def strip_generated_at(text):
        import re
        return re.sub(r'"generated_at":\s*"[^"]*"', '"generated_at":""', text)

    changed = old is None or strip_generated_at(old) != strip_generated_at(html)
    return changed


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: generate_board_status.py <xlsx_path> [output_path] [source_label]", file=sys.stderr)
        sys.exit(1)
    xlsx_arg = sys.argv[1]
    out_arg = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT
    label_arg = sys.argv[3] if len(sys.argv) > 3 else None
    changed = render(xlsx_arg, out_arg, label_arg)
    print("CHANGED" if changed else "NOCHANGE")
